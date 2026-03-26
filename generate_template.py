"""
generate_template.py
Generates the blank Ind AS 116 Excel template as a static file.
Run once; the server then serves it directly.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, datetime

OUT_DIR  = os.path.dirname(os.path.abspath(__file__))
OUT_FILE = os.path.join(OUT_DIR, "IndAS116_Blank_Template.xlsx")

# ── Colour palette ──────────────────────────────────────────────────
DARK_BLUE  = "1A2235"   # header background
GOLD       = "C9A84C"   # accent
LIGHT_GREY = "F5F5F5"   # alternate row
MID_GREY   = "DDDDDD"   # border
WHITE      = "FFFFFF"

def hdr_font(bold=True, color="FFFFFF", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def hdr_fill(hex_color=DARK_BLUE):
    return PatternFill("solid", fgColor=hex_color)

def alt_fill():
    return PatternFill("solid", fgColor=LIGHT_GREY)

def thin_border():
    s = Side(style="thin", color=MID_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ── INPUT ROWS ───────────────────────────────────────────────────────
INPUT_ROWS = [
    ("Lease Name / Asset Description",         "",           "e.g. Office Premises – Mumbai"),
    ("Lease Start Date",                        "",           "DD-MM-YYYY  e.g. 01-04-2024"),
    ("Lease End Date",                          "",           "DD-MM-YYYY  e.g. 31-03-2029"),
    ("Lease Term (months)",                     "",           "Auto-calculated or enter manually"),
    ("Lease Payment Amount (₹)",                "",           "Base periodic payment"),
    ("Payment Frequency",                       "monthly",    "monthly | quarterly | halfyearly | yearly"),
    ("Payment Timing",                          "end",        "end | beginning"),
    ("Incremental Borrowing Rate (% p.a.)",      "",           "e.g. 10.5"),
    ("Initial Direct Costs (₹)",                "0",          "Optional – legal fees, brokerage etc."),
    ("Lease Incentives Received (₹)",            "0",          "Optional – deducted from ROU asset"),
    ("Restoration / Dismantling Costs (₹)",      "0",          "Optional – added to ROU asset (Para 24(d))"),
    ("Residual Value Guarantee (₹)",             "0",          "Optional – included in last payment for PV"),
    ("Financial Year Start (Month)",             "4",          "4 = April (Indian FY)  |  1 = January"),
    ("Opening Lease Liability (₹)",              "",           "Optional – for Ind AS 116 transition"),
]

INSTR_ROWS = [
    "IND AS 116 – LEASE ACCOUNTING TOOL: UPLOAD GUIDE",
    "",
    "SHEET 1 — LEASE INPUTS",
    "  •  Fill only the VALUE column. Do NOT change the FIELD column text.",
    "  •  All dates MUST be in DD-MM-YYYY format  (e.g. 01-04-2024).",
    "  •  Payment Frequency: type exactly  monthly | quarterly | halfyearly | yearly",
    "  •  Payment Timing: type exactly  end | beginning",
    "  •  Numeric fields (amounts, rates): enter numbers only (no ₹ or % symbols).",
    "  •  Optional fields may be left blank or set to 0.",
    "",
    "SHEET 2 — PAYMENT SCHEDULE (Variable / Escalated Payments)",
    "  •  Use this sheet ONLY when lease payments differ from period to period.",
    "  •  Fill Period #, Payment Date (DD-MM-YYYY), and Payment Amount (₹).",
    "  •  Leave Payment Amount blank to use the base amount from Sheet 1.",
    "  •  Alternatively use the Escalation Clause in the tool UI to auto-generate.",
    "",
    "HOW TO UPLOAD",
    "  1.  Fill Sheets 1 and 2 as required.",
    "  2.  Save this file as .xlsx.",
    "  3.  In the tool, drag-and-drop the file on the upload area or click Browse File.",
    "  4.  The tool reads Sheet 1 for parameters and Sheet 2 for the payment schedule.",
    "  5.  Review the loaded values in the form, then click Compute.",
    "",
    "IMPORTANT NOTES",
    "  •  IBR / Incremental Borrowing Rate: obtain from your finance/treasury team.",
    "  •  Ind AS 116 requires effective interest method for lease liability amortisation.",
    "  •  ROU Asset = PV of Lease Payments + Initial Direct Costs – Incentives + Restoration.",
]


def build_inputs_sheet(wb):
    ws = wb.create_sheet("Lease Inputs")

    # ── Title rows ────────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    ws["A1"] = "IND AS 116 – LEASE ACCOUNTING INPUT TEMPLATE"
    ws["A1"].font      = Font(bold=True, color=WHITE, size=14, name="Calibri")
    ws["A1"].fill      = hdr_fill(DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    ws["A2"] = f"Generated: {datetime.date.today().strftime('%d-%m-%Y')}   |   Fill the VALUE column only. Do NOT modify the FIELD column."
    ws["A2"].font      = Font(italic=True, color="555555", size=9, name="Calibri")
    ws["A2"].fill      = hdr_fill("EAF0FB")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Blank spacer ─────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Column header row ─────────────────────────────────────────────
    headers = ["FIELD", "VALUE", "NOTES / INSTRUCTIONS"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font      = hdr_font()
        cell.fill      = hdr_fill(DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[4].height = 20

    # ── Data rows ─────────────────────────────────────────────────────
    for i, (field, default, note) in enumerate(INPUT_ROWS):
        row = i + 5
        fill = alt_fill() if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)

        fc = ws.cell(row=row, column=1, value=field)
        fc.font      = Font(bold=True, color="1A2235", size=10, name="Calibri")
        fc.fill      = fill
        fc.border    = thin_border()
        fc.alignment = Alignment(vertical="center")

        vc = ws.cell(row=row, column=2, value=default)
        vc.font      = Font(color="000000", size=10, name="Calibri")
        vc.fill      = PatternFill("solid", fgColor="FFFAEB")   # soft yellow prompt
        vc.border    = thin_border()
        vc.alignment = Alignment(vertical="center")
        # Force value column as text to prevent date serial numbers
        vc.number_format = "@"

        nc = ws.cell(row=row, column=3, value=note)
        nc.font      = Font(italic=True, color="666666", size=9, name="Calibri")
        nc.fill      = fill
        nc.border    = thin_border()
        nc.alignment = Alignment(vertical="center", wrap_text=True)

        ws.row_dimensions[row].height = 18

    # ── Column widths ─────────────────────────────────────────────────
    set_col_width(ws, 1, 44)
    set_col_width(ws, 2, 34)
    set_col_width(ws, 3, 68)

    # ── Freeze top 4 rows ────────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── Add a note for VALUE column ───────────────────────────────────
    ws.cell(row=19, column=2).comment = None   # clear any

    return ws


def build_payment_schedule_sheet(wb):
    ws = wb.create_sheet("Payment Schedule")

    # ── Instruction banner ────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    ws["A1"] = ("INSTRUCTIONS: Fill the Payment Amount for variable/escalated payments. "
                "Leave blank to use the base amount from Lease Inputs. "
                "Do NOT change the date column — dates are pre-filled by the tool.")
    ws["A1"].font      = Font(italic=True, color="555555", size=9, name="Calibri")
    ws["A1"].fill      = hdr_fill("EAF0FB")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    # ── Column headers ────────────────────────────────────────────────
    hdrs = ["PERIOD #", "PAYMENT DATE\n(DD-MM-YYYY)", "PAYMENT AMOUNT (₹)", "NOTES"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font      = hdr_font()
        cell.fill      = hdr_fill(DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[2].height = 28

    # ── 60 blank data rows (user fills) ───────────────────────────────
    for i in range(1, 61):
        row = i + 2
        fill = alt_fill() if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)

        c1 = ws.cell(row=row, column=1, value=i)
        c1.font      = Font(color="1A2235", size=10, name="Calibri")
        c1.fill      = fill; c1.border = thin_border()
        c1.alignment = Alignment(horizontal="center", vertical="center")

        c2 = ws.cell(row=row, column=2, value="")
        c2.number_format = "@"   # force text
        c2.font      = Font(color="000000", size=10, name="Calibri")
        c2.fill      = PatternFill("solid", fgColor="FFFAEB")
        c2.border    = thin_border()
        c2.alignment = Alignment(horizontal="center", vertical="center")

        c3 = ws.cell(row=row, column=3, value="")
        c3.number_format = '#,##0.00'
        c3.font      = Font(color="000000", size=10, name="Calibri")
        c3.fill      = PatternFill("solid", fgColor="FFFAEB")
        c3.border    = thin_border()
        c3.alignment = Alignment(horizontal="right", vertical="center")

        c4 = ws.cell(row=row, column=4, value="")
        c4.font      = Font(italic=True, color="888888", size=9, name="Calibri")
        c4.fill      = fill; c4.border = thin_border()

        ws.row_dimensions[row].height = 16

    set_col_width(ws, 1, 12)
    set_col_width(ws, 2, 28)
    set_col_width(ws, 3, 26)
    set_col_width(ws, 4, 44)
    ws.freeze_panes = "A3"
    return ws


def build_instructions_sheet(wb):
    ws = wb.create_sheet("Instructions")

    ws.merge_cells("A1:A1")
    ws["A1"] = "IND AS 116 – LEASE ACCOUNTING TOOL: USER GUIDE"
    ws["A1"].font      = Font(bold=True, color=WHITE, size=13, name="Calibri")
    ws["A1"].fill      = hdr_fill(DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    for i, text in enumerate(INSTR_ROWS, 2):
        cell = ws.cell(row=i, column=1, value=text)
        if text and not text.startswith(" "):
            cell.font = Font(bold=True, color=DARK_BLUE, size=10, name="Calibri")
            cell.fill = hdr_fill("EAF0FB")
        else:
            cell.font = Font(color="333333", size=10, name="Calibri")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 16

    set_col_width(ws, 1, 100)
    return ws


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default Sheet

    build_inputs_sheet(wb)
    build_payment_schedule_sheet(wb)
    build_instructions_sheet(wb)

    wb.save(OUT_FILE)
    print(f"Template saved: {OUT_FILE}")


if __name__ == "__main__":
    main()
